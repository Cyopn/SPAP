from __future__ import annotations

import hashlib
import html
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import requests

from core import storage
from core.logger import log, log_exc
from core.timezone_mx import MX_TZ, now_mx

REPORTS_DIR = Path("reports")
ALLOWED_LEVELS = ("alto", "medio", "bajo")


def _ensure_reports_dir() -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return REPORTS_DIR


def _parse_iso_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        dt = value
    else:
        txt = str(value or "").strip()
        if not txt:
            return None
        try:
            dt = datetime.fromisoformat(txt.replace("Z", "+00:00"))
        except Exception:
            try:
                dt = datetime.strptime(txt, "%Y-%m-%d")
            except Exception:
                return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=MX_TZ)
    else:
        dt = dt.astimezone(MX_TZ)
    return dt


def _parse_filter_date(value: str | None, end_of_day: bool = False) -> datetime | None:
    txt = str(value or "").strip()
    if not txt:
        return None
    try:
        dt = datetime.strptime(txt, "%Y-%m-%d").replace(tzinfo=MX_TZ)
    except Exception:
        return None
    if end_of_day:
        dt = dt + timedelta(days=1) - timedelta(microseconds=1)
    return dt


def _normalize_level(value: Any) -> str:
    v = str(value or "").strip().lower()
    if v in ALLOWED_LEVELS:
        return v
    return "medio"


def _normalize_levels(values: list[str] | tuple[str, ...] | None, default: list[str] | None = None) -> list[str]:
    source = list(values or [])
    out: list[str] = []
    for val in source:
        lvl = _normalize_level(val)
        if lvl not in out:
            out.append(lvl)
    if out:
        return out
    base = default if default is not None else ["alto", "medio", "bajo"]
    return _normalize_levels(base, default=[])


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _resolve_item_datetime(item: dict[str, Any]) -> datetime | None:
    return (
        _parse_iso_datetime(item.get("published_at"))
        or _parse_iso_datetime(item.get("extracted_at"))
        or _parse_iso_datetime(item.get("created_at"))
    )


def _build_rows(
    *,
    from_date: str | None,
    to_date: str | None,
    levels: list[str] | None,
    category: str | None,
    keyword: str | None,
    limit: int = 5000,
) -> list[dict[str, Any]]:
    items = storage.read_items(max(1, int(limit or 5000)))

    from_dt = _parse_filter_date(from_date, end_of_day=False)
    to_dt = _parse_filter_date(to_date, end_of_day=True)
    allowed_levels = set(_normalize_levels(levels))
    category_q = str(category or "").strip().lower()
    keyword_q = str(keyword or "").strip().lower()

    out: list[dict[str, Any]] = []
    for it in items:
        lvl = _normalize_level(it.get("level"))
        if allowed_levels and lvl not in allowed_levels:
            continue

        dt = _resolve_item_datetime(it)
        if from_dt and (dt is None or dt < from_dt):
            continue
        if to_dt and (dt is None or dt > to_dt):
            continue

        category_raw = str(it.get("keyword") or "").strip()
        category_l = category_raw.lower()
        if category_q and category_q not in category_l:
            continue

        title = str(it.get("title") or "").strip()
        summary = str(it.get("summary") or "").strip()
        source = str(it.get("source") or "").strip()
        url = str(it.get("url") or "").strip()
        haystack = " ".join([title, summary, category_raw, source]).lower()
        if keyword_q and keyword_q not in haystack:
            continue

        out.append(
            {
                "id": _safe_int(it.get("id"), 0),
                "date_iso": dt.isoformat() if dt else "",
                "date_text": dt.strftime("%Y-%m-%d %H:%M") if dt else "",
                "level": lvl,
                "source": source,
                "category": category_raw,
                "title": title,
                "summary": summary,
                "url": url,
                "views_count": _safe_int(it.get("views_count"), 0),
                "shares_count": _safe_int(it.get("shares_count"), 0),
            }
        )

    out.sort(key=lambda x: x.get("date_iso", ""), reverse=True)
    return out


def _build_summary(rows: list[dict[str, Any]], levels: list[str]) -> dict[str, Any]:
    dist = {"alto": 0, "medio": 0, "bajo": 0}
    for r in rows:
        lvl = _normalize_level(r.get("level"))
        dist[lvl] = dist.get(lvl, 0) + 1
    return {
        "items": len(rows),
        "views": sum(_safe_int(r.get("views_count"), 0) for r in rows),
        "shares": sum(_safe_int(r.get("shares_count"), 0) for r in rows),
        "distribution": dist,
        "levels": levels,
    }


def _infer_report_range_from_rows(rows: list[dict[str, Any]]) -> tuple[str | None, str | None]:
    dates: list[datetime] = []
    for row in rows:
        dt = _parse_iso_datetime(row.get("date_iso") or row.get("date_text"))
        if dt is not None:
            dates.append(dt)

    if not dates:
        return None, None

    return min(dates).date().isoformat(), max(dates).date().isoformat()


def _truncate(value: Any, max_len: int = 120) -> str:
    txt = str(value or "").strip().replace("\r", " ").replace("\n", " ")
    if len(txt) <= max_len:
        return txt
    return txt[: max_len - 3].rstrip() + "..."


def _resolve_logo_path(logo_ref: str | None) -> str | None:
    raw = str(logo_ref or "").strip()
    if not raw:
        return None

    if raw.lower().startswith("http://") or raw.lower().startswith("https://"):
        cache_dir = _ensure_reports_dir() / "_logo_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        key = hashlib.md5(raw.encode("utf-8")).hexdigest()
        ext = os.path.splitext(raw.split("?", 1)[0])[1] or ".img"
        target = cache_dir / f"{key}{ext}"
        if target.exists() and target.stat().st_size > 0:
            return str(target)
        try:
            resp = requests.get(raw, timeout=15)
            resp.raise_for_status()
            target.write_bytes(resp.content)
            return str(target)
        except Exception as e:
            log_exc("report: failed to download logo", e)
            return None

    p = Path(raw)
    if not p.is_absolute():
        p = Path.cwd() / p
    return str(p) if p.exists() else None


def _export_pdf_report(
    output_path: str,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
    filters: dict[str, Any],
    branding: dict[str, Any],
) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception as exc:
        raise RuntimeError(
            "No se pudo generar PDF porque falta la dependencia 'reportlab'."
        ) from exc

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitlePro",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=colors.HexColor("#0b3c5d"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "SubPro",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#334e68"),
    )
    title_cell_style = ParagraphStyle(
        "TitleCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    story = []
    logo_path = _resolve_logo_path(branding.get("logo_path"))
    logo_obj: Any = ""
    if logo_path:
        try:
            logo_obj = Image(logo_path, width=48, height=48)
        except Exception:
            logo_obj = ""

    company = _truncate(branding.get("company_name") or "SPAP", 70)
    letterhead = _truncate(branding.get("letterhead")
                           or "Reporte de monitoreo y analitica", 180)

    head_right = Paragraph(
        f"<b>{company}</b><br/>{letterhead}<br/>"
        f"Generado: {now_mx().strftime('%Y-%m-%d %H:%M UTC-6')}",
        sub_style,
    )
    head_tbl = Table([[logo_obj, head_right]], colWidths=[58, 470])
    head_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f4f8")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#bcccdc")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(head_tbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Reporte Ejecutivo de Noticias", title_style))
    story.append(Spacer(1, 8))

    summary_tbl = Table(
        [
            ["Noticias", str(summary.get("items", 0)), "Vistas",
             str(summary.get("views", 0))],
            ["Compartidos", str(summary.get("shares", 0)),
             "Niveles", ", ".join(summary.get("levels", []))],
            [
                "Distribucion",
                f"alto={summary.get('distribution', {}).get('alto', 0)} | "
                f"medio={summary.get('distribution', {}).get('medio', 0)} | "
                f"bajo={summary.get('distribution', {}).get('bajo', 0)}",
                "Rango",
                f"{filters.get('from_date') or '-'} a {filters.get('to_date') or '-'}",
            ],
        ],
        colWidths=[80, 190, 80, 190],
    )
    summary_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9e2ec")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(summary_tbl)
    story.append(Spacer(1, 10))

    filter_line = (
        f"Categoria: {filters.get('category') or 'todas'} | "
        f"Palabra clave: {filters.get('keyword') or 'ninguna'}"
    )
    story.append(Paragraph(filter_line, sub_style))
    story.append(Spacer(1, 10))

    table_data = [["Fecha", "Nivel", "Fuente",
                   "Titulo", "Vistas", "Compartidos"]]
    for row in rows:
        title_txt = str(row.get("title") or "").strip() or "(sin titulo)"
        title_cell = Paragraph(html.escape(title_txt), title_cell_style)
        table_data.append(
            [
                _truncate(row.get("date_text"), 16),
                _truncate(row.get("level"), 8),
                _truncate(row.get("source"), 20),
                title_cell,
                str(_safe_int(row.get("views_count"), 0)),
                str(_safe_int(row.get("shares_count"), 0)),
            ]
        )

    details_tbl = Table(
        table_data,
        colWidths=[72, 44, 78, 260, 45, 44],
        repeatRows=1,
    )
    details_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b3c5d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#bcccdc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    for idx in range(1, len(table_data)):
        if idx % 2 == 0:
            details_style_cmds.append(
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#f8fafc"))
            )
    details_tbl.setStyle(TableStyle(details_style_cmds))
    story.append(details_tbl)

    doc.build(story)


def _export_excel_report(
    output_path: str,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
    filters: dict[str, Any],
    branding: dict[str, Any],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        raise RuntimeError(
            "No se pudo generar Excel porque falta la dependencia 'openpyxl'."
        ) from exc

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_detail = wb.create_sheet("Detalle")

    head_fill = PatternFill(start_color="0B3C5D",
                            end_color="0B3C5D", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    label_font = Font(bold=True)

    ws_summary["A1"] = str(branding.get("company_name") or "SPAP")
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A2"] = str(branding.get("letterhead")
                           or "Reporte de monitoreo y analitica")
    ws_summary["A4"] = "Generado"
    ws_summary["B4"] = now_mx().strftime("%Y-%m-%d %H:%M UTC-6")
    ws_summary["A5"] = "Rango"
    ws_summary["B5"] = f"{filters.get('from_date') or '-'} a {filters.get('to_date') or '-'}"
    ws_summary["A6"] = "Niveles"
    ws_summary["B6"] = ", ".join(summary.get("levels", []))
    ws_summary["A7"] = "Categoria"
    ws_summary["B7"] = filters.get("category") or "todas"
    ws_summary["A8"] = "Palabra clave"
    ws_summary["B8"] = filters.get("keyword") or "ninguna"

    ws_summary["D4"] = "Noticias"
    ws_summary["E4"] = summary.get("items", 0)
    ws_summary["D5"] = "Vistas"
    ws_summary["E5"] = summary.get("views", 0)
    ws_summary["D6"] = "Compartidos"
    ws_summary["E6"] = summary.get("shares", 0)
    ws_summary["D7"] = "alto"
    ws_summary["E7"] = summary.get("distribution", {}).get("alto", 0)
    ws_summary["D8"] = "medio"
    ws_summary["E8"] = summary.get("distribution", {}).get("medio", 0)
    ws_summary["D9"] = "bajo"
    ws_summary["E9"] = summary.get("distribution", {}).get("bajo", 0)

    for ref in ("A4", "A5", "A6", "A7", "A8", "D4", "D5", "D6", "D7", "D8", "D9"):
        ws_summary[ref].font = label_font

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 44
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.column_dimensions["E"].width = 16

    headers = [
        "ID",
        "Fecha",
        "Nivel",
        "Fuente",
        "Titulo",
        "Resumen",
        "URL",
        "Vistas",
        "Compartidos",
    ]
    ws_detail.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws_detail.cell(row=1, column=col_idx)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws_detail.append(
            [
                _safe_int(row.get("id"), 0),
                row.get("date_text", ""),
                row.get("level", ""),
                row.get("source", ""),
                row.get("title", ""),
                row.get("summary", ""),
                row.get("url", ""),
                _safe_int(row.get("views_count"), 0),
                _safe_int(row.get("shares_count"), 0),
            ]
        )

    ws_detail.column_dimensions["A"].width = 8
    ws_detail.column_dimensions["B"].width = 19
    ws_detail.column_dimensions["C"].width = 12
    ws_detail.column_dimensions["D"].width = 18
    ws_detail.column_dimensions["E"].width = 55
    ws_detail.column_dimensions["F"].width = 58
    ws_detail.column_dimensions["G"].width = 48
    ws_detail.column_dimensions["H"].width = 10
    ws_detail.column_dimensions["I"].width = 12

    for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


def generate_report(
    *,
    report_format: str,
    from_date: str | None,
    to_date: str | None,
    levels: list[str] | None,
    category: str | None,
    keyword: str | None,
    branding: dict[str, Any] | None = None,
    output_dir: str | None = None,
    file_prefix: str = "reporte_noticias",
    limit: int = 5000,
) -> tuple[str, dict[str, Any]]:
    fmt = str(report_format or "pdf").strip().lower()
    if fmt not in ("pdf", "xlsx"):
        fmt = "pdf"

    rows = _build_rows(
        from_date=from_date,
        to_date=to_date,
        levels=levels,
        category=category,
        keyword=keyword,
        limit=limit,
    )
    normalized_levels = _normalize_levels(levels)
    summary = _build_summary(rows, normalized_levels)
    filters = {
        "from_date": str(from_date or "").strip() or None,
        "to_date": str(to_date or "").strip() or None,
        "category": str(category or "").strip() or None,
        "keyword": str(keyword or "").strip() or None,
    }

    auto_from_date, auto_to_date = _infer_report_range_from_rows(rows)
    if not filters.get("from_date") and auto_from_date:
        filters["from_date"] = auto_from_date
    if not filters.get("to_date") and auto_to_date:
        filters["to_date"] = auto_to_date

    report_dir = Path(output_dir) if output_dir else _ensure_reports_dir()
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = now_mx().strftime("%Y%m%d_%H%M%S")
    output_path = report_dir / f"{file_prefix}_{ts}.{fmt}"
    branding_data = branding if isinstance(branding, dict) else {}

    if fmt == "pdf":
        _export_pdf_report(
            str(output_path),
            rows=rows,
            summary=summary,
            filters=filters,
            branding=branding_data,
        )
    else:
        _export_excel_report(
            str(output_path),
            rows=rows,
            summary=summary,
            filters=filters,
            branding=branding_data,
        )

    meta = {
        "format": fmt,
        "path": str(output_path),
        "summary": summary,
        "filters": filters,
        "rows": len(rows),
    }
    return str(output_path), meta


def generate_simple_report(limit: int = 100) -> str:
    items = storage.read_items(limit)
    lines = [
        f"ID {i['id']} | {i.get('published_at') or i.get('created_at')} | {i.get('title')}"
        for i in items
    ]
    return "\n".join(lines)


if __name__ == "__main__":
    try:
        path, meta = generate_report(
            report_format="pdf",
            from_date=None,
            to_date=None,
            levels=["alto", "medio", "bajo"],
            category=None,
            keyword=None,
            branding={"company_name": "SPAP",
                      "letterhead": "Reporte automatizado"},
            file_prefix="reporte_demo",
            limit=500,
        )
        log(
            f"report: generated demo report at {path} rows={meta.get('rows')}", "INFO")
    except Exception as e:
        log_exc("report: failed to generate demo report", e)
